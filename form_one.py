import json
import string
from statistics import mean
from openpyxl import load_workbook
from pprint import pprint

# TODO Сделать инпутами, вывести в отдельынй файл
FILE: str = 'Ф1_01.xlsx'
FORM: str = 'ф 1_субъекты_рост-сниж_январь-декабрь 2023.xlsx'
REC_COL: str = 'L'
REC_COL_FORM: str = 'AK'
REC_COL_CMPR: str = 'AL'
REC_COL_MR: str = 'M'
OUTLIERS_FLAG: str = 'N'
PERIOD = '(2011 - 2019 гг. и 2022 г.)'
START_END_COll: list[int, int] = [2, 11]
START_END_ROW: list[int, int] = [10, 110]


def decode(work_book: str, sheet: str, coll: str, rows: int) -> float:
    """
    Функция для получения данных из файла EXCEL.

    Аргументы:
    wb (openpyxl.Workbook): файл EXCEL
    sheet (str): название страницы
    coll (str): столбец
    rows (int): строка


    Возвращает:
    float: список данных из указанных ячеек

    Пример использования:
    data = get_data_from_excel(wb, 'Sheet1', 1, 5, 1, 10)
    """
    sheet: str = work_book[f'{sheet}']
    value_cell: float = sheet[f'{coll}{rows}'].value
    return value_cell


def record(work_book: str, sheet: str, coll: str, rows: int,
           value: float) -> None:
    """ Функция для записи в EXEL. На вход подается:
        work_book = файл EXEL,
        sheet = страница,
        коардинаты ячейки(колонка и строчка)
        coll = колонка,
        rows = строчка
        значение = value (value имеет тип float)
    """
    sheet: str = work_book[f'{sheet}']
    sheet[f'{coll}{rows}'] = value


wb = load_workbook(filename=FILE, data_only=True)
sheet: str = wb.sheetnames
coll: list = list(string.ascii_lowercase)


def get_data(sheet: str, row: int) -> list[float]:
    data: list[float] = []

    for c in coll[START_END_COll[0] - 1:START_END_COll[1]]:
        vl = decode(wb, sheet, c, row)

        if vl == None:
            data.append(0)
        else:
            data.append(vl)

    return data


def smp(data: list) -> float:
    try:
        ls: list[float] = []
        for i in data:
            mx: float = max(data)
            mn: float = min(data)
            if i != mx and i != mn:
                ls.append(i)
        if len(ls) > 0:
            return round(sum(ls) / len(ls), 2)
        else:
            return 0
    except TypeError:
        return 0


def check_outliers(data: list):
    # Критерии для исключения выскакивающих вариант
    exclusion_criteria: dict[int, dict[int, float]] = {
        3: {95: 0.941, 99: 0.988},
        4: {95: 0.765, 99: 0.889},
        5: {95: 0.642, 99: 0.78},
        6: {95: 0.56, 99: 0.698},
        7: {95: 0.507, 99: 0.637},
        8: {95: 0.468, 99: 0.59},
        9: {95: 0.437, 99: 0.555},
        10: {95: 0.412, 99: 0.527},
        11: {95: 0.392, 99: 0.502},
        12: {95: 0.376, 99: 0.482},
        15: {95: 0.338, 99: 0.438},
        20: {95: 0.3, 99: 0.391},
        24: {95: 0.281, 99: 0.367},
        30: {95: 0.26, 99: 0.341}
    }
    # Уровень достоверности, %
    # TODO: Переименовать переменные a, b
    a: float = exclusion_criteria[len(data)][95]
    b: float = exclusion_criteria[len(data)][99]

    # Построение вариант в порядке возрастания
    data.sort()

    # Вычисление разности между наибольшей и наименьшей вариантами ряда
    diff_max_min: float = max(data) - min(data)
    outlier: set = set()

    rec_xl: dict[str, bool] = {
        'outliers_max': None,
        'outliers_min': None,
        'value': None,
    }

    try:
        ratio_max: float = (data[-1] - data[-2]) / diff_max_min
        if ratio_max > a and ratio_max > b:
            outlier.add(max(data))
            rec_xl['outliers_max'] = 'y'
            # print(f'Есть выскакивающий показатель {max(data)}')

        ratio_min: float = (data[1] - data[0]) / diff_max_min
        if ratio_min > a and ratio_min > b:
            outlier.add(min(data))
            rec_xl['outliers_min'] = 'y'
            # print(f'Есть выскакивающий показатель {min(data)}')

            # # Отладка
            # TODO Продумать отладку отдельной функцией
            # print(f'{data[1]} - {data[0]}\n------------------\n{ratio_min}\n'
            #       f'{round(ratio_min, 2)}>{a}\n{round(ratio_min, 2)}>{b}')
            # print(f'{data}\n{set(data) - outlier}')

    except ZeroDivisionError:
        rec_xl['value'] = 0

    rec_xl['value'] = round(mean(set(data) - outlier), 3)
    return rec_xl


data_for_recording = {'form_1': []}

# ЗАПИСЬ В ФАЙЛ
for sh in sheet:
    for row in range(START_END_ROW[0], START_END_ROW[1] + 1):

        data: list[float] = get_data(sh, row)
        # Записываем заголовки граф
        record(wb, sh, REC_COL, START_END_ROW[0] - 1, 'CМП (выброс макс мин)')
        record(wb, sh, REC_COL_MR, START_END_ROW[0] - 1, 'CМП (по МР)')
        record(wb, sh, OUTLIERS_FLAG, START_END_ROW[0] - 1,
               'Есть выскакивающий показатель')

        if len(data) >= 3:
            # Записываем СМП (по МР)
            record(wb, sh, f'{REC_COL_MR}', row,
                   check_outliers(data)['value'])

            data_for_recording['form_1'].append(
                {'code': sh,
                 'month': decode(wb, sh, 'A', 2).split(': ')[1],
                 'years': decode(wb, sh, 'A', 3).split(': ')[1],
                 'region': decode(wb, sh, 'A', row),
                 'value': check_outliers(data)['value'],
                 })

            # Записываем Да если есть выскакивающий показатель
            if check_outliers(data)['outliers_max'] == 'y' or check_outliers(
                    data)['outliers_min'] == 'y':
                record(wb, sh, f'{OUTLIERS_FLAG}', row, 'Да')

        if smp(data) > 0:
            # Записываем СМП (стырый способ)
            record(wb, sh, f'{REC_COL}', row, smp(data))
        else:
            record(wb, sh, f'{REC_COL}', row, 0)

wb.save(f'!{FILE}')
wb.close()


# for i in data_for_recording['form_1']:
#     print('**********')
#     for itm in i:
#         print(f'{itm} {i[itm]}')


def get_page(encoded_nosology: str, data_json_nosology: dict) -> str:
    for item in data_json_nosology:
        page: str = item['page']
        code: str = item['code']
        name: str = item['name']
        if encoded_nosology == code:
            return page


# for i in region_name_variants:
#     print(i)
none_reg = {}


def get_row_name(work_book, sheet, name):
    for row in range(1, 151):
        region_name_xl = decode(work_book, sheet, 'B', row)
        if region_name_xl == name:
            return row
        else:
            none_reg[region_name_xl] = row


#
def get_none_region(name: str, none_reg):
    region_name_variants = {
        'ЦЕНТРАЛЬНЫЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'ЦЕНТРАЛЬНЫЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'СЕВЕРО-ЗАПАДНЫЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'СЕВЕРО-ЗАПАДНЫЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'г.Санкт-Петербург': 'г. Санкт-Петербург',
        'г.Санкт-Петербург ': 'г. Санкт-Петербург',
        'г. Санкт-Петербург': 'г.Санкт-Петербург ',
        'г.Севастополь': 'г. Севастополь',
        'СЕВЕРО-КАВКАЗСКИЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'СЕВЕРО-КАВКАЗСКИЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'Республика Северная Осетия': 'Республика Северная Осетия - Алания',
        'ПРИВОЛЖСКИЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'ПРИВОЛЖСКИЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'УРАЛЬСКИЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'УРАЛЬСКИЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'СИБИРСКИЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'СИБИРСКИЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'ДАЛЬНЕВОСТОЧНЫЙ\nФЕДЕРАЛЬНЫЙ ОКРУГ': 'ДАЛЬНЕВОСТОЧНЫЙ ФЕДЕРАЛЬНЫЙ ОКРУГ',
        'Республика Саха': 'Республика Саха (Якутия)'
    }
    for key, value in region_name_variants.items():
        if name == value:
            return none_reg[key]


def compare_and_describe_changes(value1, value2) -> str:
    if value1 == 0 and value2 != 0:
        return "↑ на 100%"
    elif value1 == None:
        value1 = 0
    elif value2 == None:
        value2 = 0
    elif value1 != 0 and value2 == 0:
        return "↓ на 100%"
    elif value1 == value2:
        return "На уровне"
    else:
        g_4 = value1 / value2 if value1 > value2 else value2 / value1
        g_3 = abs(value1 * 100 / value2 - 100)
        f_4 = '↑ в' if value1 > value2 else '↓ в'
        f_3 = '↑' if value1 > value2 else '↓ на'

        if g_4 >= 1.5:
            return f"{f_4} {g_4:.1f} раза/раз"
        else:
            return f"{f_3} {g_3:.1f}%"


with open('nod_2.json', 'r', encoding='utf-8') as file:
    sheet_match = json.load(file)

match_data = sheet_match['form_1']

wb_write = load_workbook(filename=FORM, data_only=True)

for i in data_for_recording['form_1']:
    code = i['code']
    value = i['value']
    sheet_write = get_page(code, match_data)
    print(code, '|', sheet_write, '|')
    name = i['region']
    row = get_row_name(wb_write, sheet_write, name)

    record(wb_write, sheet_write, REC_COL_FORM, 6, 'CМП')
    record(wb_write, sheet_write, REC_COL_FORM, 7, PERIOD)

    if row != None:
        record(wb_write, sheet_write, REC_COL_FORM, row, value)
        value_form_one = decode(wb_write, sheet_write, 'D', row)
        record(wb_write, sheet_write, REC_COL_CMPR, row,
               compare_and_describe_changes(value1=value_form_one,
                                            value2=value))

    else:
        new_row = get_none_region(name, none_reg)
        if new_row != None:
            print(new_row)
            record(wb_write, sheet_write, REC_COL_FORM, new_row, value)
            value_form_one = decode(wb_write, sheet_write, 'D', new_row)
            record(wb_write, sheet_write, REC_COL_CMPR, new_row,
                   compare_and_describe_changes(value1=value_form_one,
                                                value2=value))

pprint(data_for_recording)

wb_write.save(f'!СМП_{FORM}')
wb_write.close()
