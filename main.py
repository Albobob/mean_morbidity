from openpyxl import load_workbook
import string
from statistics import mean

FILE = 'СМП 10 2023.xlsx'
REC_COL = 'L'
REC_COL_MR = 'M'
OUTLIERS_FLAG = 'N'
START_END_COll = [2, 10]
START_END_ROW = [11, 104]


def decode(work_book, sheet, coll, rows):
    """ Функция для чтения из EXEL. На вход подается:
 work_book = файл EXEL,
 sheet = страница,
 коардинаты ячейки(колонка и строчка)
 coll = колонка,
 rows = строчка
    """
    sheet = work_book[f'{sheet}']
    value_cell = sheet[f'{coll}{rows}'].value
    return value_cell


def record(work_book, sheet, coll, rows, value):
    """ Функция для записи в EXEL. На вход подается:
        work_book = файл EXEL,
        sheet = страница,
        коардинаты ячейки(колонка и строчка)
        coll = колонка,
        rows = строчка
        значение = value (value преобразуется в тип float)
    """
    sheet = work_book[f'{sheet}']
    sheet[f'{coll}{rows}'] = value


wb = load_workbook(filename=FILE)
sheet = wb.sheetnames

coll = list(string.ascii_lowercase)


def get_data(sheet, row):
    data = []

    for c in coll[START_END_COll[0] - 1:START_END_COll[1] + 1]:
        vl = decode(wb, sheet, c, row)
        if vl != 0:
            data.append(vl)

    return data


def smp(data):
    try:
        ls = []
        for i in data:
            mx = max(data)
            mn = min(data)
            if i != mx and i != mn:
                ls.append(i)
        if len(ls) > 0:
            return round(sum(ls) / len(ls), 2)
        else:
            return 0
    except TypeError:
        return 0


def check_outliers(data: list):
    # Критерии для исключения выскакивающих вариант
    exclusion_criteria = {
        3: {95: 0.941, 99: 0.988},
        4: {95: 0.765, 99: 0.889},
        5: {95: 0.642, 99: 0.78},
        6: {95: 0.56, 99: 0.698},
        7: {95: 0.507, 99: 0.637},
        8: {95: 0.468, 99: 0.59},
        9: {95: 0.437, 99: 0.555},
        10: {95: 0.412, 99: 0.527},
        11: {95: 0.392, 99: 0.502},
        12: {95: 0.376, 99: 0.482},
        15: {95: 0.338, 99: 0.438},
        20: {95: 0.3, 99: 0.391},
        24: {95: 0.281, 99: 0.367},
        30: {95: 0.26, 99: 0.341}
    }
    # Уровень достоверности, %
    a = exclusion_criteria[len(data)][95]
    b = exclusion_criteria[len(data)][99]

    # Построение вариант в порядке возрастания
    data.sort()
    # Вычисление разности между наибольшей и наименьшей вариантами ряда
    diff_max_min = max(data) - min(data)
    outlier = set()
    rec_xl = {
        'outliers_max': None,
        'outliers_min': None,
        'value': None,
    }
    try:
        ratio_max = (data[-1] - data[-2]) / diff_max_min
        if ratio_max > a and ratio_max > b:
            outlier.add(max(data))
            rec_xl['outliers_max'] = 'y'
            print(f'Есть выскакивающий показатель {max(data)}')

        ratio_min = (data[1] - data[0]) / diff_max_min
        if ratio_min > a and ratio_min > b:
            outlier.add(min(data))
            rec_xl['outliers_min'] = 'y'
            print(f'Есть выскакивающий показатель {min(data)}')

            # Отладка
            print(f'{data[1]} - {data[0]}\n------------------\n{ratio_min}\n'
                  f'{round(ratio_min, 2)}>{a}\n{round(ratio_min, 2)}>{b}')
            print(f'{data}\n{set(data) - outlier}')


    except ZeroDivisionError:
        rec_xl['value'] = 0

    # print(rec_xl)
    rec_xl['value'] = round(mean(set(data) - outlier), 3)
    return rec_xl


# ЗАПИСЬ В ФАЙЛ
for sh in sheet:
    for row in range(START_END_ROW[0], START_END_ROW[1] + 1):
        data = get_data(sh, row)
        record(wb, sh, REC_COL, START_END_ROW[0] - 1, 'CМП (выброс макс мин)')
        record(wb, sh, REC_COL_MR, START_END_ROW[0] - 1, 'CМП (по МР)')
        if len(data) >= 3:
            if check_outliers(data)['outliers_max'] == 'y' or check_outliers(data)['outliers_min'] == 'y':
                record(wb, sh, OUTLIERS_FLAG, START_END_ROW[0] - 1, 'Есть выскакивающий показатель')
                record(wb, sh, f'{OUTLIERS_FLAG}', row, 'Да')
            record(wb, sh, f'{REC_COL_MR}', row, check_outliers(data)['value'])
        if smp(data) > 0:
            record(wb, sh, f'{REC_COL}', row, smp(data))
        else:
            record(wb, sh, f'{REC_COL}', row, 0)

wb.save(f'!{FILE}')
